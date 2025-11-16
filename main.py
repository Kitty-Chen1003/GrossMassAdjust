import os
import sys
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QScrollArea, QSizePolicy, QMessageBox, QDesktopWidget
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from openpyxl import load_workbook


class GrossmassAdjustTool(QWidget):

    def __init__(self):

        super().__init__()
        self.save_button = None
        self.label_status = None
        self.scroll_area_label = None
        self.input_total_gross_mass = None
        self.file_path = ""
        self.df_adjusted = None
        self.adjusted_wb = None
        self.init_ui()

    def init_ui(self):

        self.setWindowTitle("GrossMass Adjust Tool")

        # 设置为屏幕大小的一半并居中
        screen = QDesktopWidget().screenGeometry()
        self.resize(screen.width() // 2, screen.height() // 2)
        frame_geo = self.frameGeometry()
        frame_geo.moveCenter(QDesktopWidget().availableGeometry().center())
        self.move(frame_geo.topLeft())

        # 外层布局（用于居中）
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setAlignment(Qt.AlignCenter)

        # 创建 QScrollArea，并设置属性
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)  # 让 scroll area 根据内容自适应
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # 横向不要滚动条（可选）
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # 纵向滚动条按需出现

        # 内层容器和布局（承载内容）
        inner_widget = QWidget()
        inner_layout = QVBoxLayout(inner_widget)
        inner_layout.setContentsMargins(100, 50, 100, 50)
        inner_layout.setSpacing(20)
        inner_layout.setAlignment(Qt.AlignTop | Qt.AlignHCenter)  # 控制内容居中靠上

        # Title
        title = QLabel("GrossMass Adjust Tool")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        inner_layout.addWidget(title)

        # flight输入框
        self.input_total_gross_mass = QLineEdit()
        self.input_total_gross_mass.setMaximumWidth(800)
        self.input_total_gross_mass.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        input_total_gross_mass_label = QLabel("total gross mass kg:")
        input_total_gross_mass_label.setObjectName("normal")

        inner_layout.addWidget(input_total_gross_mass_label)
        inner_layout.addWidget(self.input_total_gross_mass)

        # 文件上传按钮
        upload_label = QLabel("upload manifest file:")
        upload_label.setObjectName("normal")
        inner_layout.addWidget(upload_label)

        upload_button = QPushButton("Select file")
        upload_button.setFont(QFont("Microsoft YaHei", 12))
        upload_button.setMaximumWidth(800)
        upload_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        upload_button.setStyleSheet("padding: 10px;")
        upload_button.clicked.connect(self.select_file)

        # 创建 QLabel 并放入 QScrollArea 中
        self.label_status = QLabel('No file selected')
        self.label_status.setWordWrap(True)
        self.label_status.setObjectName("normal")
        self.scroll_area_label = QScrollArea()
        self.scroll_area_label.setMaximumWidth(800)
        self.scroll_area_label.setWidgetResizable(True)
        self.scroll_area_label.setMinimumHeight(100)
        self.scroll_area_label.setWidget(self.label_status)

        inner_layout.addWidget(upload_button)
        inner_layout.addSpacing(20)
        inner_layout.addWidget(self.scroll_area_label)
        inner_layout.addSpacing(20)

        # 发送按钮
        process_btn = QPushButton("Adjust GrossMassKg")
        process_btn.setObjectName("confirmbutton")
        process_btn.clicked.connect(self.adjust_grossmass)

        inner_layout.addWidget(process_btn)
        inner_layout.addSpacing(20)

        self.save_button = QPushButton("Save")
        self.save_button.setFont(QFont("Microsoft YaHei", 12))
        self.save_button.setStyleSheet("padding: 10px;")
        self.save_button.clicked.connect(self.save_adjusted)
        inner_layout.addWidget(self.save_button)
        inner_layout.addSpacing(20)

        # 底部按钮布局
        inner_layout.addStretch()

        scroll_area.setWidget(inner_widget)
        # 把内层部件加入外层居中布局
        outer_layout.addWidget(scroll_area)

    def select_file(self):

        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle('Select Manifest File')
        file_dialog.setFileMode(QFileDialog.ExistingFile)  # 只允许选择一个文件
        file_dialog.setNameFilter("Excel Files (*.xls *.xlsx)")

        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()[0]  # 只取第一个
            self.file_path = selected_file  # 保存文件路径
            self.label_status.setText(self.file_path)  # 显示文件路径

    def adjust_grossmass(self):
        if not self.file_path:
            QMessageBox.warning(self, "Error", "Please select a manifest file first.")
            return

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read file:\n{e}")
            return

        # 查找 GrossMassKg 列号（不破坏格式）
        header_row = 1
        col_index = None
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=header_row, column=col).value
            if raw is None:
                continue

            # 标准化：去空格 + 转小写
            normalized = str(raw).replace(" ", "").lower()

            if normalized == "grossmasskg":
                col_index = col
                break

        if col_index is None:
            QMessageBox.critical(
                self, "Error",
                "Column 'GrossMassKg' (case-insensitive, spaces ignored) not found."
            )
            return

        # 读取所有原值
        values = []
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_index).value

            # 如果为空则跳过（不报错）
            if cell is None or str(cell).strip() == "":
                continue

            try:
                v = float(cell)
            except:
                QMessageBox.critical(
                    self, "Invalid Value",
                    f"Row {r}: value '{cell}' cannot be converted to float."
                )
                return

            values.append(round(v, 3))

        # step 2: 计算差额
        try:
            target_total = float(self.input_total_gross_mass.text())
        except:
            QMessageBox.warning(self, "Error", "Please input a valid numeric Total GrossMassKg.")
            return

        current_total = round(sum(values), 3)
        diff = round(target_total - current_total, 3)

        n = len(values)
        if n == 0:
            QMessageBox.warning(self, "Error", "No rows found for GrossMassKg.")
            return

        step = 0.001

        min_possible_total = 0.001 * n
        if target_total < min_possible_total:
            QMessageBox.warning(
                self, "Error",
                f"Target Total GrossMassKg too small.\n"
                f"Minimum possible total for {n} rows is {min_possible_total:.3f}"
            )
            return

        # step 3: baseline 平摊
        baseline_steps = int((diff / n) / step)
        baseline = baseline_steps * step

        values_new = []
        for v in values:
            v_new = round(v + baseline, 3)
            if v_new < 0.001:
                v_new = 0.001
            values_new.append(v_new)
        values = values_new

        # step 4: 剩余差额逐行分配
        new_total = round(sum(values), 3)
        leftover = round(target_total - new_total, 3)
        leftover_steps = int(round(leftover / step))

        if leftover_steps != 0:
            sign = 1 if leftover_steps > 0 else -1
            steps = abs(leftover_steps)

            idx = 0
            while steps > 0:
                # 循环分配
                i = idx % n
                if sign < 0 and values[i] - step < 0.001:
                    idx += 1
                    continue
                values[i] = round(values[i] + sign * step, 3)
                steps -= 1
                idx += 1

        # step 5: 最终校验
        final_total = round(sum(values), 3)
        if final_total != round(target_total, 3):
            QMessageBox.warning(
                self, "Warning",
                f"Final sum mismatch!\nFinal={final_total}, Target={target_total}"
            )

        # step 6: 写回 Excel（保持所有格式不变）
        for i, v in enumerate(values, start=2):
            ws.cell(row=i, column=col_index).value = v

        self.adjusted_wb = wb
        QMessageBox.information(self, "Done", "GrossMassKg adjusted successfully!")

    def save_adjusted(self):
        if not hasattr(self, "adjusted_wb"):
            QMessageBox.warning(self, "Error", "Please adjust data before saving.")
            return

        # 只取文件名，不带路径
        original_filename = os.path.basename(self.file_path)

        # 添加前缀 Adjusted_
        default_save_name = f"Adjusted_{original_filename}"

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Adjusted File", default_save_name, "Excel Files (*.xlsx)"
        )
        if not save_path:
            return

        try:
            self.adjusted_wb.save(save_path)
            QMessageBox.information(self, "Success", f"File saved:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file:\n{e}")


def resource_path(relative):
    """获取打包后文件的真实路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath("."), relative)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    qss_path = resource_path("style.qss")
    with open(qss_path, "r", encoding="utf-8") as f:
        app.setStyleSheet(f.read())
    ui = GrossmassAdjustTool()
    ui.show()
    sys.exit(app.exec_())
